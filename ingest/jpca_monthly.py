"""Axis 2 真の需給データ — JPCA (石油化学工業協会) 月次統計 ingest.

JPCA publishes Excel time-series for:
- m2mainpd.xlsx: 主要石油化学製品生産実績 (key petrochemical production)
- ethylene equivalent trade balance, 4-resin shipment/inventory, MMA

This script:
1. Downloads the master Excel files (full history 1998-present)
2. Parses 2-row headers + year-grouped monthly data
3. Maps products → CAS via curated PRODUCT_TO_CAS table
4. Outputs parquet keyed on (year, month, product, cas)

Public data, no auth required.
"""
import io
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import requests

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "data" / "jpca"

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
HEADERS = {"User-Agent": UA}

# JPCA Excel master files — these are time-series with full history
SOURCES = {
    "main_production": {
        "url": "https://www.jpca.or.jp/files/statistics/monthly/mainpd/m2mainpd.xlsx",
        "title": "主要石油化学製品生産実績",
        "unit": "千トン",
        "header_rows": [3, 4],  # 2-row header (general name / detail)
        "data_start_row": 9,
    },
}

# Map JPCA product display name → CAS (1 CAS per product; polymers use generic CAS)
PRODUCT_TO_CAS: dict[str, str] = {
    "エチレン": "74-85-1",
    "低密度ポリエチレン": "9002-88-4",
    "高密度ポリエチレン": "9002-88-4",  # PE generic
    "ポリプロピレン": "9003-07-0",
    "スチレンモノマー": "100-42-5",
    "スチレンポリマー": "9003-53-6",  # actually polystyrene per JPCA's quirky 2-row header
    "ポリスチレン": "9003-53-6",
    "塩ビモノマー": "75-01-4",
    "塩ビポリマー": "9002-86-2",
    "モノマー": "100-42-5",  # JPCA header parsing concat artifact — first "モノマー" col = SM
    "MMAモノマー": "80-62-6",  # methyl methacrylate
    "エチレンオキサイド": "75-21-8",
    "エチレングリコール": "107-21-1",
    "アセトアルデヒド": "75-07-0",
    "アクリロニトリル": "107-13-1",
    "ＳＢＲ（ソリッド）": "9003-55-8",
    "ＳＢＲ\n（ソリッド）": "9003-55-8",
    "合成ゴムＳＢＲ\n（ソリッド）": "9003-55-8",  # header concat
    "合成ゴムＳＢＲ（ソリッド）": "9003-55-8",
    "ＢＲ（ソリッド）": "9003-17-2",
    "ＢＲ\n（ソリッド）": "9003-17-2",
    "ベンゼン": "71-43-2",
    "芳香族ベンゼン": "71-43-2",  # header concat
    "トルエン": "108-88-3",
    "キシレン": "1330-20-7",
}

# Patterns that mark non-data rows
SKIP_LABEL_PATTERNS = [
    r"^（?前.*比[）)]?$",
    r"^\[前.*比\]$",
    r"^［前.*比］$",
    r"^\(前.*比\)$",
]


def fetch_excel(url: str) -> bytes:
    r = requests.get(url, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return r.content


def parse_headers(ws, header_rows: list[int]) -> list[str]:
    """Combine 2-row headers into single product names per column."""
    row_a = [c.value for c in ws[header_rows[0]]]
    row_b = [c.value for c in ws[header_rows[1]]]
    out = []
    for a, b in zip(row_a, row_b):
        a_s = str(a).strip() if a else ""
        b_s = str(b).strip() if b else ""
        if a_s and b_s:
            # Join, handling cases like "低密度ポリ" + "エチレン" → "低密度ポリエチレン"
            combined = a_s + b_s if not a_s.endswith(b_s[:1]) else a_s + b_s
            out.append(combined)
        else:
            out.append(a_s or b_s)
    return out


def is_skip_label(label: str) -> bool:
    if not label:
        return True
    for pat in SKIP_LABEL_PATTERNS:
        if re.match(pat, label):
            return True
    # Skip quarterly/annual sums (e.g., "1-3月計", "1-12月計")
    if re.search(r"\d+-\d+月計", label):
        return True
    return False


YEAR_MONTH_RE = re.compile(r"(\d{4})年\s*(\d{1,2})月")
MONTH_ONLY_RE = re.compile(r"^\s*(\d{1,2})月\s*$")


def parse_label_to_period(label: str, current_year: int | None) -> tuple[int | None, int | None]:
    """Return (year, month) or (None, None) if not a monthly data row."""
    if not label:
        return None, None
    m = YEAR_MONTH_RE.search(label)
    if m:
        return int(m.group(1)), int(m.group(2))
    m2 = MONTH_ONLY_RE.match(label)
    if m2 and current_year:
        return current_year, int(m2.group(1))
    return None, None


def to_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in ("-", "—", "—"):
        return None
    # Remove brackets / parens
    s = re.sub(r"[（）()\[\]［］〔〕]", "", s)
    s = s.replace("▲", "-").replace("△", "-")
    try:
        return float(s)
    except ValueError:
        return None


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fetched_at = datetime.now(timezone.utc).isoformat()
    all_rows: list[dict] = []

    for source_key, meta in SOURCES.items():
        print(f"=== {source_key}: {meta['url']} ===")
        try:
            data = fetch_excel(meta["url"])
        except requests.RequestException as e:
            print(f"  FAIL: {e}")
            continue
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        print(f"  Sheet '{ws.title}', {ws.max_row}r x {ws.max_column}c")

        # Find header rows by looking for "エチレン" / known product names in row range
        header_rows = meta["header_rows"]
        products = parse_headers(ws, header_rows)
        print(f"  Products: {[p[:20] for p in products if p][:15]}")

        # Iterate data rows
        current_year: int | None = None
        rows_kept = 0
        for ri in range(meta["data_start_row"], ws.max_row + 1):
            label_cell = ws.cell(row=ri, column=2).value  # col B
            label = str(label_cell).strip() if label_cell else ""
            if is_skip_label(label):
                continue
            yr, mo = parse_label_to_period(label, current_year)
            if yr is None:
                continue
            current_year = yr  # update context for subsequent "X月" rows
            # Read all product columns (start from col C = index 3)
            for col_idx, product_name in enumerate(products[2:], start=3):
                if not product_name or product_name in ("None", "-"):
                    continue
                v = to_float(ws.cell(row=ri, column=col_idx).value)
                if v is None:
                    continue
                cas = PRODUCT_TO_CAS.get(product_name)
                all_rows.append({
                    "source": source_key,
                    "year": yr,
                    "month": mo,
                    "period": f"{yr}-{mo:02d}",
                    "product": product_name,
                    "cas": cas,
                    "value": v,
                    "unit": meta["unit"],
                    "_fetched_at": fetched_at,
                })
            rows_kept += 1
        print(f"  Kept {rows_kept} monthly rows")

    if not all_rows:
        print("No data collected")
        return

    df = pd.DataFrame(all_rows)
    # Sanity check
    cas_resolved = df["cas"].notna().sum()
    products_unique = df["product"].nunique()
    print(f"\nTotal rows: {len(df):,}")
    print(f"Unique products: {products_unique}")
    print(f"Rows with CAS resolved: {cas_resolved:,} ({100*cas_resolved/len(df):.0f}%)")
    print(f"Period range: {df['period'].min()} → {df['period'].max()}")
    print(f"\nProducts × CAS mapping:")
    print(df.drop_duplicates("product")[["product", "cas"]].to_string(index=False))

    stamp = datetime.now().strftime("%Y%m%d")
    out_path = OUT_DIR / f"jpca_monthly_{stamp}.parquet"
    df.to_parquet(out_path, index=False, compression="snappy")
    print(f"\nWrote {len(df)} rows → {out_path}")


if __name__ == "__main__":
    main()
